"""
exports.py — brand-styled Excel and PDF generation.

Excel: a Summary sheet (one row per booking) + an Itemized sheet (every line
with code, net, VAT amount, gross) — what accounting needs.
PDF: an Adina letterhead price sheet; single bookings also get the itemised table.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from core import net_from_gross, fmt_eur, short_name

TERRA = "B65A33"
INK = "2B2622"
SAGE = "6F7D6A"


# ---------------------------------------------------------------------------
# helpers to normalise a record (works for both live Booking objects and
# dict rows loaded from the database)
# ---------------------------------------------------------------------------
def _rec(b):
    if isinstance(b, dict):
        return {
            "booking_id": b["booking_id"], "customer": b["customer"],
            "qty": b["qty"], "unit": b["unit"], "package": b["package"],
            "event_date": b["event_date"], "net": b["net"], "vat7": b["vat7"],
            "vat19": b["vat19"], "gross": b["gross"], "adjust": b["adjust"],
            "adjust_note": b["adjust_note"], "total": b["total"],
            "lines": json.loads(b["lines_json"]),
            "services": json.loads(b["services_json"]),
        }
    return {
        "booking_id": b.booking_id, "customer": b.customer, "qty": b.qty,
        "unit": b.unit, "package": b.package, "event_date": b.event_date,
        "net": b.net, "vat7": b.vat7, "vat19": b.vat19, "gross": b.gross,
        "adjust": b.adjust, "adjust_note": b.adjust_note, "total": b.total,
        "lines": b.lines, "services": b.services,
    }


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------
def export_excel(bookings, path):
    recs = [_rec(b) for b in bookings]
    wb = Workbook()

    head_fill = PatternFill("solid", fgColor=TERRA)
    head_font = Font(color="FFFFFF", bold=True, size=10)
    total_fill = PatternFill("solid", fgColor=INK)
    total_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E8DED3")
    border = Border(bottom=thin)
    right = Alignment(horizontal="right")

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    headers = ["Booking ID", "Customer", "Qty", "Unit", "Package", "Event date",
               "Add-ons", "Net", "VAT 7%", "VAT 19%", "Gross",
               "Adjustment", "Adj. note", "Final total"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font

    for r in recs:
        addons = ", ".join(f"{n} (€{p:g})" for n, p, _ in r["services"]) or "—"
        ws.append([
            r["booking_id"], r["customer"], r["qty"], r["unit"], r["package"],
            r["event_date"], addons, r["net"], r["vat7"], r["vat19"],
            r["gross"], r["adjust"], r["adjust_note"], r["total"],
        ])

    if len(recs) > 1:
        ws.append([])
        t = {k: round(sum(x[k] for x in recs), 2)
             for k in ("net", "vat7", "vat19", "gross", "adjust", "total")}
        ws.append(["", "GRAND TOTAL", "", "", "", "", "",
                   t["net"], t["vat7"], t["vat19"], t["gross"],
                   t["adjust"], "", t["total"]])
        last = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=last, column=c).fill = total_fill
            ws.cell(row=last, column=c).font = total_font

    widths = [16, 20, 6, 7, 34, 13, 26, 11, 10, 10, 11, 12, 16, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in "HIJKLN":
        for cell in ws[col]:
            cell.alignment = right

    # --- Itemized sheet ---
    ws2 = wb.create_sheet("Itemized")
    h2 = ["Booking ID", "Customer", "Code", "Line item", "Qty",
          "VAT %", "Net", "VAT amount", "Gross"]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        ws2.cell(row=1, column=c).fill = head_fill
        ws2.cell(row=1, column=c).font = head_font

    for r in recs:
        for name, ug, rate, code in r["lines"]:
            g = ug * r["qty"]
            n = net_from_gross(g, rate)
            ws2.append([r["booking_id"], r["customer"], code, name, r["qty"],
                        rate, round(n, 2), round(g - n, 2), round(g, 2)])
        for name, ug, rate in r["services"]:
            g = ug * r["qty"]
            n = net_from_gross(g, rate)
            ws2.append([r["booking_id"], r["customer"], "SVC", name, r["qty"],
                        rate, round(n, 2), round(g - n, 2), round(g, 2)])

    for i, w in enumerate([16, 20, 9, 28, 6, 7, 10, 11, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def export_pdf(bookings, path):
    recs = [_rec(b) for b in bookings]
    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    elems = []

    logo = ParagraphStyle("logo", parent=styles["Title"], fontName="Times-Bold",
                          fontSize=22, textColor=colors.HexColor("#" + INK),
                          spaceAfter=0)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=8,
                         textColor=colors.grey, spaceBefore=2)
    elems.append(Paragraph(f'Adina<font color="#{TERRA}">.</font>', logo))
    elems.append(Paragraph("MEETINGS &amp; EVENTS &nbsp;·&nbsp; BERLIN MITTE &nbsp;—&nbsp; Price Sheet", sub))
    elems.append(Spacer(1, 8))

    head = ["ID", "Customer", "Qty", "Package", "Add-ons",
            "Net", "VAT 7%", "VAT 19%", "Adjust", "Total"]
    data = [head]
    for r in recs:
        addons = ", ".join(n for n, _, _ in r["services"]) or "—"
        adj = "—" if r["adjust"] == 0 else (("−" if r["adjust"] < 0 else "+") +
                                            fmt_eur(abs(r["adjust"]))[1:])
        data.append([r["booking_id"], r["customer"], str(r["qty"]),
                     short_name(r["package"]), addons,
                     fmt_eur(r["net"]), fmt_eur(r["vat7"]), fmt_eur(r["vat19"]),
                     adj, fmt_eur(r["total"])])

    t = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + TERRA)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (5, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF7F3")]),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#" + TERRA)),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if len(recs) > 1:
        grand = fmt_eur(round(sum(r["total"] for r in recs), 2))
        data.append(["", "", "", "", "", "", "", "", "Grand total", grand])
        last = len(data) - 1
        style += [
            ("BACKGROUND", (0, last), (-1, last), colors.HexColor("#" + INK)),
            ("TEXTCOLOR", (0, last), (-1, last), colors.white),
            ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
        ]
        t = Table(data, repeatRows=1)
    t.setStyle(TableStyle(style))
    elems.append(t)

    # single booking -> itemised detail
    if len(recs) == 1:
        r = recs[0]
        elems.append(Spacer(1, 10))
        d2 = [["Code", "Line item", "VAT", "Net", "VAT amt", "Gross"]]
        for name, ug, rate, code in r["lines"]:
            g = ug * r["qty"]
            n = net_from_gross(g, rate)
            d2.append([code, name, f"{rate}%", fmt_eur(n), fmt_eur(g - n), fmt_eur(g)])
        for name, ug, rate in r["services"]:
            g = ug * r["qty"]
            n = net_from_gross(g, rate)
            d2.append(["SVC", name, f"{rate}%", fmt_eur(n), fmt_eur(g - n), fmt_eur(g)])
        t2 = Table(d2, repeatRows=1, colWidths=[22*mm, 70*mm, 16*mm, 28*mm, 28*mm, 28*mm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + SAGE)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF7F3")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(t2)

    elems.append(Spacer(1, 14))
    foot = ParagraphStyle("foot", parent=styles["Normal"], fontSize=7.5,
                          textColor=colors.grey)
    elems.append(Paragraph(
        "Adina Apartment Hotel Berlin Mitte · Platz vor dem Neuen Tor 6, "
        "10115 Berlin · Alto Restaurant &amp; Bar", foot))

    doc.build(elems)
    return path
